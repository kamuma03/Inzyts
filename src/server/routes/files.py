from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
import shutil
import uuid
from pathlib import Path
from typing import List
from werkzeug.utils import secure_filename as werkzeug_secure_filename
from src.config import settings
from src.server.middleware.auth import verify_token
from src.utils.logger import get_logger
from src.utils.path_validator import validate_path_within, ensure_dir
from src.server.models.schemas import FilePreview, DBTestRequest, DBTestResponse, SQLPreviewRequest, APIPreviewRequest
import magic

logger = get_logger()

router = APIRouter(prefix="/files", tags=["files"])

# Resolve once at module load so the path is stable regardless of CWD at request
# time. Mirrors the pattern used in analysis.py so path-safety checks in both
# modules refer to the same absolute directory.
UPLOAD_DIR = str(ensure_dir(settings.upload_dir_resolved))

MAX_UPLOAD_SIZE = 100 * 1024 * 1024  # 100MB

ALLOWED_MIMES = [
    # CSV
    "text/csv",
    "text/plain",
    "application/csv",
    "application/x-csv",
    "text/x-csv",
    "text/comma-separated-values",
    "text/x-comma-separated-values",
    # Parquet
    "application/vnd.apache.parquet",
    # NOTE: application/octet-stream intentionally excluded — too permissive.
    # Parquet/Excel files detected as octet-stream are handled by the
    # extension-based fallback in _validate_upload().
    # Excel
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
    # JSON
    "application/json",
    "text/json",
    # Empty files (edge case)
    "application/x-empty",
    "inode/x-empty",
]



def _validate_upload(file_size: int, header: bytes, filename: str) -> None:
    if len(header) == 0:
        logger.warning(f"Rejected empty file upload: {filename}")
        raise HTTPException(status_code=400, detail=f"Empty file not allowed: {filename}")

    if file_size > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File {filename} too large ({file_size} bytes). Maximum allowed: {MAX_UPLOAD_SIZE} bytes",
        )

    mime = magic.Magic(mime=True)
    mime_type = mime.from_buffer(header)

    if mime_type not in ALLOWED_MIMES:
        # Extension-based fallbacks for formats whose magic bytes don't match
        # the canonical MIME type (e.g. .parquet → octet-stream, .xlsx → zip).
        extension_ok = (
            (filename.endswith(".parquet") and mime_type == "application/octet-stream")
            or (filename.endswith((".xlsx", ".xls")) and mime_type in ("application/zip", "application/x-ole-storage"))
        )
        if not extension_ok:
            logger.warning(f"Rejected file {filename} with mime type {mime_type}")
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type for {filename}: {mime_type}. Allowed: CSV, Parquet, Excel (.xlsx/.xls), and JSON.",
            )

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...), _token: str = Depends(verify_token)
):
    """
    Upload a CSV/Parquet file to the server with MIME type validation.
    """
    try:
        filename = file.filename or "unknown"
        # Read first 2048 bytes for MIME detection
        header = await file.read(2048)
        await file.seek(0)

        if len(header) == 0:
            logger.warning(f"Rejected empty file upload: {filename}")
            raise HTTPException(status_code=400, detail="Empty files are not allowed")

        # Check file size
        file.file.seek(0, 2)
        file_size = file.file.tell()
        file.file.seek(0)
        
        _validate_upload(file_size, header, filename)

        # Generate unique filename to prevent overwrite clashes
        # Sanitize filename to prevent path traversal
        safe_filename = werkzeug_secure_filename(filename)
        file_ext = Path(safe_filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = Path(UPLOAD_DIR) / unique_filename

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        return {
            "filename": file.filename,
            "saved_path": unique_filename,  # Return only relative filename
            "size": file_path.stat().st_size,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"File upload failed: {e}")
        raise HTTPException(
            status_code=500, detail="Internal server error during file upload"
        )





@router.post("/upload_batch")
async def upload_batch(
    files: List[UploadFile] = File(...), _token: str = Depends(verify_token)
):
    """
    Upload multiple CSV/Parquet files with MIME type validation.
    """
    results = []
    try:
        for file in files:
            filename = file.filename or "unknown"
            # MIME validation (mirrors single upload)
            header = await file.read(2048)
            await file.seek(0)

            if len(header) == 0:
                logger.warning(f"Rejected empty file in batch: {filename}")
                raise HTTPException(
                    status_code=400, detail=f"Empty file not allowed: {filename}"
                )

            # Check file size
            file.file.seek(0, 2)
            file_size = file.file.tell()
            file.file.seek(0)
            
            _validate_upload(file_size, header, filename)

            # Sanitize filename and generate unique path
            safe_filename = werkzeug_secure_filename(filename)
            file_ext = Path(safe_filename).suffix
            unique_filename = f"{uuid.uuid4()}{file_ext}"
            file_path = Path(UPLOAD_DIR) / unique_filename

            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)

            results.append(
                {
                    "filename": filename,
                    "saved_path": unique_filename,
                    "size": file_path.stat().st_size,
                }
            )
        return results
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Batch upload failed: {e}")
        raise HTTPException(
            status_code=500, detail="Internal server error during batch upload"
        )


@router.get("/preview", response_model=FilePreview)
async def preview_file(path: str, _token: str = Depends(verify_token)):
    """
    Get a preview of the file content.

    Reads the first 5 rows of the file to allow the user to confirm the content
    and columns before starting analysis.

    Supports: .csv, .parquet, .log

    Args:
        path: Absolute path to the file on the server.

    Returns:
        FilePreview: Structure containing columns, rows (first 5), and total count.

    Raises:
        HTTPException: If file is not found or cannot be read.
    """
    import pandas as pd

    # Security check: Ensure we are only reading from the upload directory
    path_abs = validate_path_within(
        path,
        [Path(UPLOAD_DIR)],
        resolve_relative_to=Path(UPLOAD_DIR),
        must_exist=True,
        error_label="file",
    )
    path = str(path_abs)  # Use resolved absolute path for rest of logic

    try:
        ext = Path(path).suffix.lower()

        if ext == ".parquet":
            import pyarrow.parquet as pq
            pf = pq.ParquetFile(path)
            total_rows = pf.metadata.num_rows
            # Read only the first row group, then slice to 5 rows
            first_batch = next(pf.iter_batches(batch_size=5))
            df_head = first_batch.to_pandas()
            columns = df_head.columns.tolist()
            rows = df_head.head(5).to_dict("records")
        elif ext in (".xlsx", ".xls"):
            df_head = pd.read_excel(path, nrows=5)
            # For total row count, use openpyxl without loading all data
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                total_rows = ws.max_row - 1 if ws.max_row else 0  # subtract header
                wb.close()
            except Exception:
                total_rows = len(df_head)
            columns = df_head.columns.tolist()
            rows = df_head.to_dict("records")
        elif ext == ".json":
            # Read JSON — for large files this is unavoidable, but JSON previews
            # are typically small API responses.
            df = pd.read_json(path)
            total_rows = len(df)
            columns = df.columns.tolist()
            rows = df.head(5).to_dict("records")
        else:
            # Robust CSV/Log handling using our utility
            from src.utils.file_utils import load_csv_robust

            # First, count total rows efficiently (without parsing)
            def count_lines(filepath):
                with open(filepath, "rb") as f:
                    return sum(1 for _ in f)

            # Estimate total rows (can be slow for huge files, but safe)
            # For huge files, maybe we skip exact count?
            # Existing code did this, so we keep it.
            try:
                total_rows = count_lines(path) - 1  # Subtract header
            except Exception:
                total_rows = -1  # validation failed or binary

            # Load only first 5 rows for preview using robust loader
            # We pass nrows=5 to read_csv internally
            df_head = load_csv_robust(path, nrows=5)

            columns = df_head.columns.tolist()
            rows = df_head.to_dict("records")

        return FilePreview(
            filename=Path(path).name,
            columns=columns,
            rows=rows,
            total_rows=max(0, total_rows),
        )
    except Exception as e:
        logger.error(f"File preview failed for {path}: {e}")
        raise HTTPException(status_code=400, detail="Failed to read file preview")


@router.post("/db-test", response_model=DBTestResponse)
async def test_db_connection(
    req: DBTestRequest, _token: str = Depends(verify_token)
):
    """
    Test a database connection and return basic info (dialect, host, table list).

    Uses a 10-second connect timeout to avoid blocking on unreachable hosts.
    """
    from urllib.parse import urlparse
    from sqlalchemy import create_engine, inspect

    # db_uri scheme already validated by DBTestRequest Pydantic model.
    parsed = urlparse(req.db_uri)
    try:
        engine = create_engine(
            req.db_uri,
            pool_timeout=10,
            pool_pre_ping=True,
        )
        with engine.connect() as conn:
            inspector = inspect(engine)
            tables = inspector.get_table_names()[:50]  # Cap to avoid huge response
        engine.dispose()

        return DBTestResponse(
            status="ok",
            dialect=parsed.scheme,
            host=parsed.hostname,
            tables=tables,
        )
    except Exception as e:
        logger.error(f"DB connection test failed: {e}")
        # Sanitize: only show exception type and generic message
        return DBTestResponse(
            status="error",
            dialect=parsed.scheme,
            host=parsed.hostname,
            error=f"Connection failed: {type(e).__name__}",
        )


@router.post("/sql-preview", response_model=FilePreview)
async def preview_sql_query(
    req: SQLPreviewRequest, _token: str = Depends(verify_token)
):
    """
    Execute a SQL query and return the first 5 rows as a preview.

    Validates the query is SELECT-only before execution.
    """
    import pandas as pd
    from sqlalchemy import create_engine, text

    # db_uri scheme already validated by SQLPreviewRequest Pydantic model.

    # Validate SELECT-only (user-provided query — must check before execution)
    from src.agents.sql_agent import _validate_select_only
    validation_error = _validate_select_only(req.query)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    try:
        engine = create_engine(req.db_uri, pool_timeout=10, pool_pre_ping=True)
        with engine.connect() as conn:
            chunk_iter = pd.read_sql(text(req.query), conn, chunksize=6)
            chunk = next(iter(chunk_iter))
            chunk_iter.close()

        rows = chunk.head(5).to_dict("records")
        columns = chunk.columns.tolist()

        engine.dispose()
        return FilePreview(
            filename="sql_preview",
            columns=columns,
            rows=rows,
            total_rows=len(chunk),
        )
    except StopIteration:
        return FilePreview(filename="sql_preview", columns=[], rows=[], total_rows=0)
    except Exception as e:
        logger.error(f"SQL preview failed: {e}")
        raise HTTPException(status_code=400, detail=f"SQL preview failed: {type(e).__name__}")


@router.post("/api-preview", response_model=FilePreview)
async def preview_api_endpoint(
    req: APIPreviewRequest, _token: str = Depends(verify_token)
):
    """
    Fetch a single page from a REST API and return the first 5 rows as a preview.
    """
    import requests as http_requests
    import pandas as pd
    from src.agents.api_agent import _build_auth_headers, _extract_data_with_jmespath, _is_private_ip

    if _is_private_ip(req.api_url):
        raise HTTPException(status_code=400, detail="API URL resolves to a private/reserved IP.")

    headers = dict(req.api_headers or {})
    headers.update(_build_auth_headers(req.api_auth))
    headers.setdefault("Accept", "application/json")

    try:
        response = http_requests.get(req.api_url, headers=headers, timeout=10)
        response.raise_for_status()
        data = response.json()
        records = _extract_data_with_jmespath(data, req.json_path)

        if not records:
            return FilePreview(filename="api_preview", columns=[], rows=[], total_rows=0)

        df = pd.json_normalize(records)
        return FilePreview(
            filename="api_preview",
            columns=df.columns.tolist(),
            rows=df.head(5).to_dict("records"),
            total_rows=len(df),
        )
    except http_requests.exceptions.Timeout:
        raise HTTPException(status_code=400, detail="API request timed out")
    except http_requests.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else "unknown"
        raise HTTPException(status_code=400, detail=f"API returned HTTP {status}")
    except Exception as e:
        logger.error(f"API preview failed: {e}")
        raise HTTPException(status_code=400, detail=f"API preview failed: {type(e).__name__}")
